'''
The _equilibrate module implements a python interface between the Equilibrate
and EquilibrateState objective-C classes (and the subclasses of equilibrate).

Any function defined below that is specific to one of the subclasses if Equilibrtae,
namely EquilibrateUsingMELTSv102, EquilibrateUsingMELTSv110. EquilibrateUsingMELTSv120,
EquilibrateUsingpMELTSv561, EquilibrateUsingMELTSwithDEW, or EquilibrateUsingStixrude,
are so indicated.  Otherwise all functions are built to interact with the Equilibrate
class and the EquilibrateState class.
'''

from thermoengine.core import *

import numpy as np
import scipy as sp

import xml.etree.ElementTree as ET
import locale

# Objective-C imports
import ctypes
from ctypes import cdll
from ctypes import util
from collections import OrderedDict
from rubicon.objc import ObjCClass, NSObject, objc_method
cdll.LoadLibrary(util.find_library('phaseobjc'))

# Excel imports
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

__all__ = ['MELTSmodel']

class MELTSmodel:
	"""Class for creating an instance of the Equilibrate PhaseObjC class that is tailored to
	calculate equilibrium phase assemblages using one of the MELTS calibrations.

	Valid initializers are version='1.0.2', '1.1.0', '1.2.0', '5.6.1'

	"""
	def __init__(self, version='1.0.2'):
		if version == '1.0.2':
			MELTSclass = ObjCClass('EquilibrateUsingMELTSv102')
		elif version == '1.1.0':
			MELTSclass = ObjCClass('EquilibrateUsingMELTSv110')
		elif version == '1.2.0':
			MELTSclass = ObjCClass('EquilibrateUsingMELTSv120')
		elif version == '5.6.1':
			MELTSclass = ObjCClass('EquilibrateUsingpMELTSv561')
		else:
			assert False, 'Unknown version of MELTS stipulated'
		self.melts = MELTSclass.alloc().init()
		self.melts.setDebugS_(0)
		self.melts.setDebugV_(0)

		oxide_names_NSArray = self.melts.oxideNames()
		self.noxides = oxide_names_NSArray.count
		self.oxide_names_a = [oxide_names_NSArray.objectAtIndex_(i) for i in range(self.noxides)]

		phase_names_NSArray = self.melts.phaseNames()
		self.nphases = phase_names_NSArray.count
		self.phase_names_a = [phase_names_NSArray.objectAtIndex_(i) for i in range(self.nphases)]

		locale.setlocale(locale.LC_NUMERIC, '')

	def print_default_settings(self):
		"""Prints list of options and tolerance settings for algorithm.

		"""
		option_settings = {
			'PPOPTIONS_DISTRIBUTION':     ('Make failure in element redistribution non-fatal',       self.melts.PPOPTIONS_DISTRIBUTION.boolValue),
			'PPOPTIONS_CHEM_POTENTIAL':   ('Make chemical potential recast errors non-fatal',        self.melts.PPOPTIONS_CHEM_POTENTIAL.boolValue),
			'PPOPTIONS_RANK':             ('Make rank deficient quadratic solutions fatal',          self.melts.PPOPTIONS_RANK.boolValue),
			'PPOPTIONS_ZERO_LINEAR':      ('Make "zero" steplength linear searches non-fatal',       self.melts.PPOPTIONS_ZERO_LINEAR.boolValue),
			'PPOPTIONS_QUAD_ITERS':       ('Terminate if quadratic iterations ever exceed maximum',  self.melts.PPOPTIONS_QUAD_ITERS.boolValue),
			'PPOPTIONS_FORCE_COMPONENTS': ('Never zero a component concentration in a system phase', self.melts.PPOPTIONS_FORCE_COMPONENTS.boolValue),
			'PPOPTIONS_DETECT_MINIMAL':   ('Prevent minimal energy test for convergence',            self.melts.PPOPTIONS_DETECT_MINIMAL.boolValue)
		}
		for key in option_settings:
			value = option_settings[key]
			print ("{0:<60.60s} {1:<30.30s} {2:1d}".format(value[0], key, value[1]))

		tolerance_settings = {
			'PPPARAMETERS_CHEM_POTENTIAL': \
			('Residual tolerance for acceptable recasting of phase chemical potentials to those of the elements', \
				float(self.melts.PPPARAMETERS_CHEM_POTENTIAL)),
			'PPPARAMETERS_RANK': \
			('Tolerance (relative to Hessian norm) for computation of pseudorank during quadratic search', \
				self.melts.PPPARAMETERS_RANK),
			'PPPARAMETERS_QUAD_OPTIMAL': \
			('Residual norm relative tolerance for optimal quadratic convergence', \
				self.melts.PPPARAMETERS_QUAD_OPTIMAL),
			'PPPARAMETERS_QUAD_SUBOPTIMAL': \
			('Residual norm relative tolerance for sub-optimal quadratic convergence', \
				self.melts.PPPARAMETERS_QUAD_SUBOPTIMAL),
			'PPPARAMETERS_QUAD_MAX_ITERS': \
			('Maximal number of quadratic iterations', \
				self.melts.PPPARAMETERS_QUAD_MAX_ITERS),
			'PPPARAMETERS_LINEAR_MAX_ITERS': \
			('Maximal number of linear search iterations', \
				self.melts.PPPARAMETERS_LINEAR_MAX_ITERS),
			'PPPARAMETERS_LINEAR_RELTEST': \
			('Relative convergence criteria for estimation of the minimum during a linear search step', \
				self.melts.PPPARAMETERS_LINEAR_RELTEST),
			'PPPARAMETERS_LINEAR_MIN_STEPLENGTH': \
			('Minimum allowed steplength in a linear search step', \
				self.melts.PPPARAMETERS_LINEAR_MIN_STEPLENGTH),
			'PPPARAMETERS_COMPONENT_MINIMUM': \
			('Minimum molar concentration of a component in a phase (absolute value)', \
				self.melts.PPPARAMETERS_COMPONENT_MINIMUM),
			'PPPARAMETERS_MASSOUT': \
			('Phase mass that triggers removal of that phase from the system', \
				self.melts.PPPARAMETERS_MASSOUT),
			'PPPARAMETERS_MOLESIN': \
			('Moles of phase added to system on detection of phase saturation', \
				self.melts.PPPARAMETERS_MOLESIN),
			'PPPARAMETERS_LINEAR_MINIMAL': \
			('Number of quadratic iterations over which to average potential for minimal energy convergence test', \
				self.melts.PPPARAMETERS_LINEAR_MINIMAL),
			'PPPARAMETERS_CYCLES_MAX': \
			('Maximal number of phase inclusion/phase removal cycles permitted', \
				self.melts.PPPARAMETERS_CYCLES_MAX)
		}
		for key in tolerance_settings:
			value = tolerance_settings[key]
			print ("{0:<70.70s} {1:<30.30s} {2:13.6e}".format(value[0], key, value[1]))

	def set_debug_state(self, debugS=False, debugV=False):
		"""Sets debug output level for Equilibrate class and subclasses.

		Parameters
		==========
		debugS : boolean, optional
			Sets on or off low level debug output. Default is off (False).
		debugV : boolean, optional
			Sets on or off high level debug output. Default is off (False).

		"""
		if debugS == False:
			self.melts.setDebugS_(0)
		else:
			self.melts.setDebugS_(1)
		if debugV == False:
			self.melts.setDebugV_(0)
		else:
			self.melts.setDebugV_(1)

	def get_oxide_names(self):
		"""Retrieves a list of system oxides.

		Composition of the system can be expressed in terms of these oxides.

		Returns
		-------
		array : strings
		    An array of strings naming system components in terms of oxides

		"""
		return self.oxide_names_a

	def get_phase_names(self):
		"""Retrieves a list of system phases.

		Names of phases known to the system.

		Returns
		-------
		array : strings
		    An array of strings naming system phases.

		"""
		return self.phase_names_a

	def get_phase_inclusion_status(self):
		"""Retrieves a dictionary of the inclusion status of each phase.

		Returns
		-------
		dict : dictionary
		    A dictionary of boolean values indicating the inclusion status of each phase (key) known to the system.

		"""
		dict = {}
		state_NSdict = self.melts.getPermissablePhasesState()
		for phase in self.phase_names_a:
			value = state_NSdict.valueForKey_(phase).boolValue
			if value == 1:
				value = True
			else:
				value = False
			dict[phase] = value
		return dict

	def set_phase_inclusion_status(self, status_d):
		"""Sets the inclusion status of specified phases.

		Parameters
		----------
		status_d : dictionary
			A dictionary of phase name keys and bollean values. True sets inclusion and False prevents inclusion of a phase
			in the equilibrium assemblage.  Note that the chemical affinity of the phase will still be calculated even
			if the inclusion level is set to False.
		"""
		state_NSdict = self.melts.getPermissablePhasesState()
		nsarray_cls = ObjCClass('NSMutableArray')
		nsarray = nsarray_cls.arrayWithCapacity_(self.nphases)
		for phase in self.phase_names_a:
			value = state_NSdict.valueForKey_(phase).boolValue
			if phase in status_d:
				if status_d[phase] == True:
					value = 1
				else:
					value = 0
			nsarray.addObject_(ObjCClass('NSNumber').numberWithBool_(value))
		self.melts.setPermissablePhasesState_(nsarray)


	def set_bulk_composition(self, oxide_d={}):
		"""Sets the bulk composition of the system.

		This function first tests if the composition is feasible before setting the bulk
		composition of the system.  You should check to make sure the composition is
		feasible before proceeding.

		Parameters
		----------
		oxide_d : a python dictionary
			A dictionary of oxide names and values, e.g. {'SiO2':77.8, 'Al2O3':12.0, ..., 'H2O':3.74}

		Returns
		-------
		boolean : True or False
		    True if the composition is feasible, in which case the composition of the system is defined.
		    False if the composition is infeasible, in which case the composition of the system is undefined.

		Notes
		-----
		Feasibility call has yet to be implemented: (Objective-C method call:)
		-(BOOL)compositionIsFeasible:(NSArray \*)compositionInWtPercentOxides forSolution:(id <SolutionPhaseProtocol>)omniComponentPhase;

		"""
		wt = (ctypes.c_double*self.noxides)()
		ctypes.cast(wt, ctypes.POINTER(ctypes.c_double))
		for i in range(0, self.noxides):
			wt[i] = 0.0
		for key,value in oxide_d.items():
			index = self.oxide_names_a.index(key)
			wt[index] = value
		self.melts.setComposition_(wt)

	def set_temperature(self, t_c=800.0):
		"""Sets the temperature of the system and reinitializes a calculation sequence.

		Parameters
		----------
		t_c : float optional
			Float value to set the system temperature in degrees centigrade. Default is 800 °C.

		"""
		self.t = t_c + 273.15
		self.melts.setTemperature_(self.t)

	def set_entropy(self, s):
		"""Sets the entropy of the system and reinitializes a calculation sequence.

		Parameters
		----------
		s : float
			Float value to set the system entropy in J/K

		"""
		self.entropy = s
		self.melts.setEntropy_(self.entropy)

	def set_pressure(self, p_mpa=200.0):
		"""Sets the pressure of the system and reinitializes a calculation sequence.

		Parameters
		----------
		p_mpa : float optional
			Float value to set the system pressure in mega-Pascals, default is 200 MPa

		"""
		self.p = p_mpa*10.0
		self.melts.setPressure_(self.p)

	def set_volume(self, v):
		"""Sets the volume of the system and reinitializes a calculation sequence.

		Parameters
		----------
		v : float
			Float value to set the system volume in J/bar

		"""
		self.volume = v
		self.melts.setVolume_(self.volume)

	def equilibrate_tp(self, T_a, P_a, initialize=False):
		"""Determines the equilibrium phase assemblage at a temperature-pressure point or along a series of T-P points.

		The bulk composition of the system must first be set by calling the function:
		set_bulk_composition

		Parameters
		----------
		t_a : float or numpy array of floats
			Temperature in degrees centigrade.  Either a scaler values or a numpy array of float values must be provided.
		p_a : float or numpy array of floats
			Pressure in mega-Pascals. Either a scaler values or a numpy array of float values must be provided.
			NOTE: If both ``t_a`` and ``p_a`` are arrays, then they must both be the same length.
		initialize : bool, optional
			True if this is a T-, P-point that starts a sequence of calculations.

			False if this is a continuation T-,P-pair or series of pairs.

		Returns
		-------
		output_a : an array of tuples
			tuple = (status, T, P, xmlout):
			status is a string indicating the status of the calculation: success/failiure, Reason for success/failure.
			T is a float value corresponding to the temperature in degrees centigrade.
			P is a float value corresponding to the pressure in mega-Pascals.
			xmlout is an xml document tree of the type xml.etree.ElementTree.  The xml tree contains information on the masses and
			abundances of all phases in the system.  ``xmlout`` is utilized as input for a number of functions in this package that
			retrieve properties of the equilibrium assemblage.

		Notes
		-----
		The ``xmlout`` document tree will be expanded to include thermodynamic properties of the phases and chemical affinities of
		phases not present in the equilibrium assemblage.

		"""
		T_a, P_a = fill_array(T_a, P_a)
		output_a = []
		for ind,(T,P) in enumerate(zip(T_a,P_a)):
			if initialize:
				self.melts.setTemperature_(T+273.15)
				self.melts.setPressure_(P*10.0)

				nsarray_cls = ObjCClass('NSMutableArray')
				nsarraykeys = nsarray_cls.arrayWithCapacity_(5)
				nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('ordinate'))
				nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('abscissa'))
				nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('imposeBuffer'))
				nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('buffer'))
				nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('bufferOffset'))

				nsarrayvalues = nsarray_cls.arrayWithCapacity_(5)
				nsarrayvalues.addObject_(ObjCClass('NSString').stringWithString_('Temperature (°C)'))
				nsarrayvalues.addObject_(ObjCClass('NSString').stringWithString_('Pressure (MPa)'))
				nsarrayvalues.addObject_(ObjCClass('NSNumber').numberWithBool_(0))
				nsarrayvalues.addObject_(ObjCClass('NSString').stringWithString_('QFM'))
				nsarrayvalues.addObject_(ObjCClass('NSNumber').numberWithBool(0))  # Double initialization does not work (?)

				nsdict_cls = ObjCClass('NSDictionary')
				nsdict = nsdict_cls.dictionaryWithObjects_forKeys_(nsarrayvalues, nsarraykeys)
				self.melts.setCalculationOptions_(nsdict)

			else:
				self.melts.incrementTemperature_(T+273.15)
				self.melts.incrementPressure_(P*10.0)
			output_NSDictionary =  self.melts.execute()
			xmlout = ET.fromstring(self.melts.equilibrateResultsAsXML())
			output_a.append((output_NSDictionary.objectForKey_('status'), T, P, xmlout))
		return output_a

	def equilibrate_sp(self, S_a, P_a, initialize=False):
		"""Determines the equilibrium phase assemblage at an entropy-pressure point or along a series of S-P points.

		The bulk composition of the system must first be set by calling the function:
		``set_bulk_composition``

		Parameters
		----------
		S_a : float or numpy array of floats
			Entropy in Joules per Kelvins.  Either a scaler values or a numpy array of float values must be provided.
		P_a : float or numpy array of floats
			Pressure in mega-Pascals. Either a scaler values or a numpy array of float values must be provided.
			NOTE: If both ``t_a`` and ``p_a`` are arrays, then they must both be the same length.
		initialize : bool, optional
			True if this is a S-, P-point that starts a sequence of calculations.

			False if this is a continuation S-,P-pair or series of pairs.

		Returns
		-------
		output_a : an array of tuples
			tuple = (status, T, P, xmlout):
			status is a string indicating the status of the calculation: success/failiure, Reason for success/failure.
			T is a float value corresponding to the temperature in degrees centigrade.
			P is a float value corresponding to the pressure in mega-Pascals.
			xmlout is an xml document tree of the type xml.etree.ElementTree.  The xml tree contains information on the masses and
			abundances of all phases in the system.  ``xmlout`` is utilized as input for a number of functions in this package that
			retrieve properties of the equilibrium assemblage.

		Notes
		-----
		The ``xmlout`` document tree will be expanded to include thermodynamic properties of the phases and chemical affinities of
		phases not present in the equilibrium assemblage.

		"""
		S_a, P_a = fill_array(S_a, P_a)
		output_a = []
		for ind,(S,P) in enumerate(zip(S_a,P_a)):
			if initialize:
				self.melts.setEntropy_(S)
				self.melts.setPressure_(P*10.0)

				nsarray_cls = ObjCClass('NSMutableArray')
				nsarraykeys = nsarray_cls.arrayWithCapacity_(5)
				nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('ordinate'))
				nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('abscissa'))
				nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('imposeBuffer'))
				nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('buffer'))
				nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('bufferOffset'))

				nsarrayvalues = nsarray_cls.arrayWithCapacity_(5)
				nsarrayvalues.addObject_(ObjCClass('NSString').stringWithString_('Entropy (J/K-kg)'))
				nsarrayvalues.addObject_(ObjCClass('NSString').stringWithString_('Pressure (MPa)'))
				nsarrayvalues.addObject_(ObjCClass('NSNumber').numberWithBool_(0))
				nsarrayvalues.addObject_(ObjCClass('NSString').stringWithString_('QFM'))
				nsarrayvalues.addObject_(ObjCClass('NSNumber').numberWithBool(0))  # Double initialization does not work (?)

				nsdict_cls = ObjCClass('NSDictionary')
				nsdict = nsdict_cls.dictionaryWithObjects_forKeys_(nsarrayvalues, nsarraykeys)
				self.melts.setCalculationOptions_(nsdict)

			else:
				self.melts.incrementEntropy_(S)
				self.melts.incrementPressure_(P*10.0)
			output_NSDictionary =  self.melts.execute()
			xmlout = ET.fromstring(self.melts.equilibrateResultsAsXML())
			T = locale.atof(xmlout.find(".//Temperature").text)
			output_a.append((output_NSDictionary.objectForKey_('status'), T, P, xmlout))
		return output_a

	def equilibrate_tv(self, T_a, V_a, initialize=False):
		"""Determines the equilibrium phase assemblage at a temperature-volume point
		or along a series of T-V points.

		The bulk composition of the system must first be set by calling the function:
		set_bulk_composition

		Parameters
		----------
		T_a : float or numpy array of floats
			Temperature in degrees centigrade.  Either a scaler values or a numpy array of float values must be provided.
		V_a : float or numpy array of floats
			Volume in Joules per bar. Either a scaler values or a numpy array of float values must be provided.
			NOTE: If both ``t_a`` and ``p_a`` are arrays, then they must both be the same length.
		initialize : bool, optional
			True if this is a T-, V-point that starts a sequence of calculations.

			False if this is a continuation T-,V-pair or series of pairs.

		Returns
		-------
		output_a : an array of tuples
			tuple = (status, T, P, xmlout):
			status is a string indicating the status of the calculation: success/failiure, Reason for success/failure.
			T is a float value corresponding to the temperature in degrees centigrade.
			P is a float value correcponding to the pressure in mega-Pascals.
			xmlout is an xml document tree of the type xml.etree.ElementTree.  The xml tree contains information on the masses and
			abundances of all phases in the system.  ``xmlout`` is utilized as input for a number of functions in this package that
			retrieve properties of the equilibrium assemblage.

		Notes
		-----
		The ``xmlout`` document tree will be expanded to include thermodynamic properties of the phases and chemical affinities of
		phases not present in the equilibrium assemblage.

		"""
		T_a, V_a = fill_array(T_a, V_a)
		output_a = []
		for ind,(T,V) in enumerate(zip(T_a,V_a)):
			if initialize:
				self.melts.setTemperature_(T+273.15)
				self.melts.setVolume_(V)

				nsarray_cls = ObjCClass('NSMutableArray')
				nsarraykeys = nsarray_cls.arrayWithCapacity_(5)
				nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('ordinate'))
				nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('abscissa'))
				nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('imposeBuffer'))
				nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('buffer'))
				nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('bufferOffset'))

				nsarrayvalues = nsarray_cls.arrayWithCapacity_(5)
				nsarrayvalues.addObject_(ObjCClass('NSString').stringWithString_('Temperature (°C)'))
				nsarrayvalues.addObject_(ObjCClass('NSString').stringWithString_('Volume (cc/kg)'))
				nsarrayvalues.addObject_(ObjCClass('NSNumber').numberWithBool_(0))
				nsarrayvalues.addObject_(ObjCClass('NSString').stringWithString_('QFM'))
				nsarrayvalues.addObject_(ObjCClass('NSNumber').numberWithBool(0))  # Double initialization does not work (?)

				nsdict_cls = ObjCClass('NSDictionary')
				nsdict = nsdict_cls.dictionaryWithObjects_forKeys_(nsarrayvalues, nsarraykeys)
				self.melts.setCalculationOptions_(nsdict)

			else:
				self.melts.incrementTemperature_(T+273.15)
				self.melts.incrementVolume_(V)
			output_NSDictionary =  self.melts.execute()
			xmlout = ET.fromstring(self.melts.equilibrateResultsAsXML())
			output_a.append((output_NSDictionary.objectForKey_('status'), T, P, xmlout))
		return output_a

	def equilibrate_sv(self, S_a, V_a, initialize=False):
		"""Determines the equilibrium phase assemblage at an entropy-volume point
		or along a series of S-V points.

		The bulk composition of the system must first be set by calling the function:
		set_bulk_composition

		Parameters
		----------
		S_a : float or numpy array of floats
			Entropy in Joules per Kelvins.  Either a scaler values or a numpy array of float values must be provided.
		V_a : float or numpy array of floats
			Volume in Joules per bar. Either a scaler values or a numpy array of float values must be provided.
			NOTE: If both ``t_a`` and ``p_a`` are arrays, then they must both be the same length.
		initialize : bool, optional
			True if this is a S-, V-point that starts a sequence of calculations.

			False if this is a continuation S-,V-pair or series of pairs.

		Returns
		-------
		output_a : an array of tuples
			tuple = (status, T, P, xmlout):
			status is a string indicating the status of the calculation: success/failiure, Reason for success/failure.
			T is a float value corresponding to the temperature in degrees centigrade.
			P is a float value corresponding to the pressure in mega-Pascals.
			xmlout is an xml document tree of the type xml.etree.ElementTree.  The xml tree contains information on the masses and
			abundances of all phases in the system.  ``xmlout`` is utilized as input for a number of functions in this package that
			retrieve properties of the equilibrium assemblage.

		Notes
		-----
		The ``xmlout`` document tree will be expanded to include thermodynamic properties of the phases and chemical affinities of
		phases not present in the equilibrium assemblage.

		"""
		S_a, V_a = fill_array(S_a, V_a)
		output_a = []
		for ind,(S,V) in enumerate(zip(S_a,V_a)):
			if initialize:
				self.melts.setEntropy_(S)
				self.melts.setVolume_(V)

				nsarray_cls = ObjCClass('NSMutableArray')
				nsarraykeys = nsarray_cls.arrayWithCapacity_(5)
				nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('ordinate'))
				nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('abscissa'))
				nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('imposeBuffer'))
				nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('buffer'))
				nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('bufferOffset'))

				nsarrayvalues = nsarray_cls.arrayWithCapacity_(5)
				nsarrayvalues.addObject_(ObjCClass('NSString').stringWithString_('Entropy (J/K-kg)'))
				nsarrayvalues.addObject_(ObjCClass('NSString').stringWithString_('Volume (cc/kg)'))
				nsarrayvalues.addObject_(ObjCClass('NSNumber').numberWithBool_(0))
				nsarrayvalues.addObject_(ObjCClass('NSString').stringWithString_('QFM'))
				nsarrayvalues.addObject_(ObjCClass('NSNumber').numberWithBool(0))  # Double initialization does not work (?)

				nsdict_cls = ObjCClass('NSDictionary')
				nsdict = nsdict_cls.dictionaryWithObjects_forKeys_(nsarrayvalues, nsarraykeys)
				self.melts.setCalculationOptions_(nsdict)

			else:
				self.melts.incrementEntropy_(S)
				self.melts.incrementVolume_(V)
			output_NSDictionary =  self.melts.execute()
			xmlout = ET.fromstring(self.melts.equilibrateResultsAsXML())
			output_a.append((output_NSDictionary.objectForKey_('status'), T, P, xmlout))
		return output_a

	def get_list_of_phases_in_assemblage(self, root):
		"""Returns a list of phase in the specified equilibrium assemblage.

		Parameters
		----------
		root : type xml.etree.ElementTree
			An xml document tree returned by the function equilibrate_xx

		Returns
		-------
		list : list
			A Python list of all phases in the equilibrium assemblage

		"""
		list = []
		for phase in self.phase_names_a:
			if root.find(".//System/Phase[@Type='" + phase + "']/Mass") != None:
				list.append(phase)
		return list

	def get_dictionary_of_default_fractionation_coefficients(self, fracLiq=False, fracSolid=True, fracFluid=True, fracCoeff=1.0):
		"""Returns a dictionary of default coefficients for phase fractionation.

		These coefficients may be modified by the user. They are used when setting
		the extent to which a phase will fractionate from the equilibrium assemblage.

		Parameters
		----------
		fracLiq : bool, optional
			Flag to indicate if liquid phases should be fractionated from the system.
			Default is False.
		fracSolids : bool, optional
			Flag to indicate if solid phases should be fractionated from the system.
			Default is True.
		fracFluids : bool, optional
			Flag to indicate if fluid phases should be fractionated from the system.
			Default is True.
		fracCoeff : float, optional
			Fractionation coefficient, which gives the fractional extend to which the
			mass of the phase is extracted during phase fractionation.
			Default is 1.0 (100%).

		Returns
		-------
		dict : dictionary
			A Python dictionary keyed on phases with values giving the extent (in fractional
			units) that a phase will fractionation mass

		Notes
		-----
		This dictionary is provided as input to the function fractionate_phases().
		The default configuration fractionates all solid/fluid phases and retains liquid.

		"""
		dict ={}
		for phase in self.phase_names_a:
			if phase == 'Liquid':
				if fracLiq:
					dict[phase] = fracCoeff
				else:
					dict[phase] = 0.0
			elif phase == 'Water' or phase == 'Fluid':
				if fracFluid:
					dict[phase] = fracCoeff
				else:
					dict[phase] = 0.00
			else:
				if fracSolid:
					dict[phase] = fracCoeff
				else:
					dict[phase] = 0.0
		return dict

	def get_mass_of_phase(self, root, phase_name='System'):
		"""Returns the mass of a phase in the specified equilibrium assemblage.

		Parameters
		----------
		root : type xml.etree.ElementTree
			An xml document tree returned by the function equilibrate_xx
		phase_name : string, optional
			The name of the phase who abundance is requested, or the string 'System', which returns the
			combined mass of all phases in the system.  Default value is 'System'

		Returns
		-------
		value : float
			The mass of the phase in the equilibrium assemblage specified by ``root``, in grams.
			If the specified phase is not in the equilibrium assemblage, a value of zero is retuned.

		"""
		if phase_name == 'System':
			value = 0.0
			for phase in self.phase_names_a:
				if root.find(".//System/Phase[@Type='" + phase + "']/Mass") != None:
					value += float(root.find(".//System/Phase[@Type='" + phase + "']/Mass").text)
		else:
			if root.find(".//System/Phase[@Type='" + phase_name + "']/Mass") != None:
				value = float(root.find(".//System/Phase[@Type='" + phase_name + "']/Mass").text)
			else:
				value = 0.0
		return value

	def get_composition_of_phase(self, root, phase_name='System', mode='oxide_wt'):
		"""Returns the composition of a phase in the specified equilibrium assemblage
		as a dictionary, with composition tabluated in the specified mode.

		Parameters
		----------
		root : type xml.etree.ElementTree
			An xml document tree returned by the function equilibrate_xx
		phase_name : string, optional
			The name of the phase who abundance is requested, or the string 'System', which returns the
			combined mass of all phases in the system.  Default value is 'System'
		mode : string, optional
			Controls the contents of the returned dictionary.
			'formula' returns a dictionary with the string 'formula' as key and value set to a string representation of the phase
			formula. For pure compoonent phases this is the standard phase formula.  For solutions, this is the actual formula
			constructed by weighting the endmember components by their mole fractions.
			'oxide_wt' returns a dictionary of oxide string keys with values in wt%. This is a valid ``mode`` for all ``phase_name``
			entries.
			'component' returns a dictionary of endmember component keys with values in mole fraction. The length of this dictionary
			will vary dependening on the number of components that describe the solution. Pure phases return an empty dictionary, as does
			``phase_name`` set to 'System'
			The default value of ``mode`` is 'oxide_wt'

		Returns
		-------
		dict : dictionary
			A dictionary describing the composition of ``phase_name`` according to the ``mode`` specified. The
			dictionary will be empty if ``phase_name`` is not present in the equilibrium assemblage.  It will
			also be empty for certain cases described above under ``mode``

		"""
		dict = {}
		if phase_name == 'System':
			if mode == 'oxide_wt':
				oxides = list(root.findall(".//Composition/Oxide"))
				for oxide in oxides:
					key = oxide.attrib['Type']
					value = float(oxide.text)
					dict[key] = value
		else:
			phase = root.find(".//System/Phase[@Type='" + phase_name + "']")
			if phase != None:
				if mode == 'formula':
					dict['formula'] = phase.find("Formula").text
				elif mode == 'oxide_wt':
					oxides = list(phase.findall("Oxide"))
					for oxide in oxides:
						key = oxide.attrib['Type']
						value = float(oxide.text)
						dict[key] = value
				elif mode == 'component':
					components = list(phase.findall("Component"))
					if not components:
						dict['formula'] = phase.find("Formula").text
					else:
						for component in components:
							key = component.attrib['Name']
							value = float(component.text)
							dict[key] = value
		return dict

	def fractionate_phases(self, root, frac_coeff):
		"""Fractionates phases from the system.
		Partitions and maintains an internal dictionary of fractionates and automatically modifies system bulk
		composition to reflect fractionation.

		Parameters
		----------
		root : type xml.etree.ElementTree
			An xml document tree returned by the function equilibrate_xx
		frac_coeff : dictionary
			A dictionary keyed on phase names with values that indicate the fraction of each phase that should
			fractionate.
			See get_dictionary_of_default_fractionation_coefficients()

		Returns
		-------
		dict : dictionary
			A dictionary keyed on phase names with values corresponding to a dictionary of phase properties.
			Keys are property names.
		"""
		dict = {}

		bc = {}
		bcFactor = locale.atof(root.find(".//Mass").text)/100.0
		oxides = list(root.findall(".//Composition/Oxide"))
		for oxide in oxides:
			key = oxide.attrib['Type']
			value = locale.atof(oxide.text)
			bc[key] = value

		phases = list(root.findall(".//System/Phase"))
		for phase in phases:
			phase_name = phase.attrib['Type']
			dict[phase_name] = {}
			fraction = frac_coeff[phase_name]
			if fraction > 0.0:
				mass = locale.atof(phase.find("Mass").text)
				oxides = list(phase.findall("Oxide"))
				for oxide in oxides:
					key = oxide.attrib['Type']
					value = locale.atof(oxide.text)
					dict[phase_name][key] = value*fraction*mass/100.0
					bc[key] -= value*fraction*mass/100.0
					if bc[key] < 0.0:
						bc[key] = 0.0
						print ("Zeroed: " + key)

		self.set_bulk_composition(bc)
		return dict

	def get_thermo_properties_of_phase_components(self, root, phase_name, mode='mu'):
		"""Returns a dictionary of the specified component thermodynamic properties of the designated phase.

		Parameters
		----------
		root : type xml.etree.ElementTree
			An xml document tree returned by the function equilibrate_xx
		phase_name : string
			The name of the phase who abundance is requested.
		mode : string, optional
			Controls the contents of the returned dictionary.
			'mu' returns a dictionary of endmember component keys with values of chemical potential. The length of this dictionary
			will vary depending on the number of components that describe the solution. Pure phases return a dictionary of unit length,
			with the phase name as key and their specific Gibbs Free energy as value (J/g).
			'excess' returns a dictionary of endmember component keys with values of excess chemical potential.  The length of this dictionary
			will vary depending on the number of components that describe the solution. Pure phases return a dictionary of unit length,
			with the phase name as key and zero as value.
			'activity' returns a dictionary of endmember component keys with values of component activity.  The length of this dictionary
			will vary depending on the number of components that describe the solution. Pure phases return a dictionary of unit length,
			with the phase name as key and unity as value.

		Returns
		-------
		dict : dictionary
			A dictionary describing the thermodynamic properties of components in ``phase_name`` according to the ``mode`` specified. The
			dictionary will be empty if ``phase_name`` is not present in the equilibrium assemblage.

		"""
		dict = {}
		phase = root.find(".//System/Phase[@Type='" + phase_name + "']")
		if phase != None:
			if mode == 'mu':
				mus = list(phase.findall("ChemicalPotential"))
				if not mus:
					value = locale.atof(phase.find("GibbsFreeEnergy").text)
					mass = float(phase.find("Mass").text)
					dict[phase_name] = value/mass
				else:
					for mu in mus:
						key = mu.attrib['Name']
						value = locale.atof(mu.text)
						dict[key] = value

			elif mode == 'excess':
				mus = list(phase.findall("ExcessChemicalPotential"))
				if not mus:
					dict[phase_name] = 0.0
				else:
					for mu in mus:
						key = mu.attrib['Name']
						value = locale.atof(mu.text)
						dict[key] = value

			elif mode == 'activity':
				mus = list(phase.findall("ExcessChemicalPotential"))
				if not mus:
					dict[phase_name] = 1.0
				else:
					t = locale.atof(root.find(".//Temperature").text) + 273.15
					for mu in mus:
						key = mu.attrib['Name']
						value = locale.atof(mu.text)
						dict[key] = np.exp(value/8.3143/t)

		return dict

	def get_list_of_properties(self):
		"""Returns a list of properties reported for each phase in an equilibrium assemblage.

		Returns
		-------
		list : list
			A Python list of all properties of phases in an equilibrium assemblage

		"""
		list = ['Mass','GibbsFreeEnergy','Enthalpy','Entropy','HeatCapacity','DcpDt', 'Volume','DvDt','DvDp','D2vDt2','D2vDtDp','D2vDp2', \
		        'Density','Alpha','Beta','K',"K'",'Gamma']
		return list

	def get_units_of_property(self, prop='Mass'):
		"""Returns the units of a specified property.

		Returns
		-------
		string : string
			The units of the specified property.  Returns 'none' if property is invalid.

		"""
		dict = {'Mass':'g','GibbsFreeEnergy':'J','Enthalpy':'J','Entropy':'J/K','HeatCapacity':'J/K','DcpDt':'J/K^2', \
		        'Volume':'J/bar','DvDt':'J/bar-K','DvDp':'J/bar^2','D2vDt2':'J/bar-K^2','D2vDtDp':'J/bar^2-K','D2vDp2':'J/bar^3', \
		        'Density':'g/cm^3','Alpha':'1/K','Beta':'1/bar','K':'GPa',"K'":'none','Gamma':'none'}
		return dict.get(prop)

	def get_property_of_phase(self, root, phase_name='System', property_name='Mass'):
		"""Returns the specified property of a phase in the specified equilibrium assemblage.

		Parameters
		----------
		root : type xml.etree.ElementTree
			An xml document tree returned by the function equilibrate_xx
		phase_name : string, optional
			The name of the phase who property is requested, or the string 'System', which returns the
			combined property of all phases in the system.  Default value is 'System'
		property_name : string, optional
			The name of the property to be returned. Default value is 'Mass'

		Returns
		-------
		value : float
			The property of the phase in the equilibrium assemblage specified by ``root``, in standard units.
			If the specified phase is not in the equilibrium assemblage, a value of zero is returned.
			If the property is not in the standard list, a value of zero is returned.

		"""
		standard = ['Mass','GibbsFreeEnergy','Enthalpy','Entropy','HeatCapacity','DcpDt', 'Volume','DvDt','DvDp','D2vDt2','D2vDtDp','D2vDp2']
		derived = ['Density','Alpha','Beta','K',"K'",'Gamma']

		if property_name in standard:
			if phase_name == 'System':
				value = 0.0
				for phase in self.phase_names_a:
					if root.find(".//System/Phase[@Type='" + phase + "']/" + property_name) != None:
						value += locale.atof(root.find(".//System/Phase[@Type='" + phase + "']/" + property_name).text)
			else:
				if root.find(".//System/Phase[@Type='" + phase_name + "']/" + property_name) != None:
					value = locale.atof(root.find(".//System/Phase[@Type='" + phase_name + "']/" + property_name).text)
				else:
					value = 0.0
		elif property_name in derived:
			if property_name == 'Density':
				if phase_name == 'System':
					volume = 0.0
					mass = 0.0
					for phase in self.phase_names_a:
						if root.find(".//System/Phase[@Type='" + phase + "']/Mass") != None:
							volume += float(root.find(".//System/Phase[@Type='" + phase + "']/Volume").text)
							mass += float(root.find(".//System/Phase[@Type='" + phase + "']/Mass").text)
					value = mass/volume/10.0 # g/cc
				else:
					if root.find(".//System/Phase[@Type='" + phase_name + "']/Mass") != None:
						volume = float(root.find(".//System/Phase[@Type='" + phase_name + "']/Volume").text)
						mass = float(root.find(".//System/Phase[@Type='" + phase_name + "']/Mass").text)
						value = mass/volume/10.0 # g/cc
					else:
						value = 0.0
			elif property_name == 'Alpha':
				if phase_name == 'System':
					volume = 0.0
					dvdt = 0.0
					for phase in self.phase_names_a:
						if root.find(".//System/Phase[@Type='" + phase + "']/Mass") != None:
							volume += float(root.find(".//System/Phase[@Type='" + phase + "']/Volume").text)
							dvdt += float(root.find(".//System/Phase[@Type='" + phase + "']/DvDt").text)
					value = dvdt/volume
				else:
					if root.find(".//System/Phase[@Type='" + phase_name + "']/Mass") != None:
						volume = float(root.find(".//System/Phase[@Type='" + phase_name + "']/Volume").text)
						dvdt = float(root.find(".//System/Phase[@Type='" + phase_name + "']/DvDt").text)
						value = dvdt/volume
					else:
						value = 0.0
			elif property_name == 'Beta':
				if phase_name == 'System':
					volume = 0.0
					dvdp = 0.0
					for phase in self.phase_names_a:
						if root.find(".//System/Phase[@Type='" + phase + "']/Mass") != None:
							volume += float(root.find(".//System/Phase[@Type='" + phase + "']/Volume").text)
							dvdp += float(root.find(".//System/Phase[@Type='" + phase + "']/DvDp").text)
					value = -dvdp/volume
				else:
					if root.find(".//System/Phase[@Type='" + phase_name + "']/Mass") != None:
						volume = float(root.find(".//System/Phase[@Type='" + phase_name + "']/Volume").text)
						dvdp = float(root.find(".//System/Phase[@Type='" + phase_name + "']/DvDp").text)
						value = -dvdp/volume
					else:
						value = 0.0
			elif property_name == 'K':
				if phase_name == 'System':
					volume = 0.0
					dvdp = 0.0
					for phase in self.phase_names_a:
						if root.find(".//System/Phase[@Type='" + phase + "']/Mass") != None:
							volume += float(root.find(".//System/Phase[@Type='" + phase + "']/Volume").text)
							dvdp += float(root.find(".//System/Phase[@Type='" + phase + "']/DvDp").text)
					value = -volume/dvdp/10000.0
				else:
					if root.find(".//System/Phase[@Type='" + phase_name + "']/Mass") != None:
						volume = float(root.find(".//System/Phase[@Type='" + phase_name + "']/Volume").text)
						dvdp = float(root.find(".//System/Phase[@Type='" + phase_name + "']/DvDp").text)
						value = -volume/dvdp/10000.0
					else:
						value = 0.0
			elif property_name == "K'":
				if phase_name == 'System':
					volume = 0.0
					dvdp = 0.0
					d2vdp2 = 0.0
					for phase in self.phase_names_a:
						if root.find(".//System/Phase[@Type='" + phase + "']/Mass") != None:
							volume += float(root.find(".//System/Phase[@Type='" + phase + "']/Volume").text)
							dvdp += float(root.find(".//System/Phase[@Type='" + phase + "']/DvDp").text)
							d2vdp2 += float(root.find(".//System/Phase[@Type='" + phase + "']/D2vDp2").text)
					value = (volume*d2vdp2/dvdp/dvdp-1.0)
				else:
					if root.find(".//System/Phase[@Type='" + phase_name + "']/Mass") != None:
						volume = float(root.find(".//System/Phase[@Type='" + phase_name + "']/Volume").text)
						dvdp = float(root.find(".//System/Phase[@Type='" + phase_name + "']/DvDp").text)
						d2vdp2 = float(root.find(".//System/Phase[@Type='" + phase_name + "']/D2vDp2").text)
						value = (volume*d2vdp2/dvdp/dvdp-1.0)
					else:
						value = 0.0
			elif property_name == 'Gamma':
				dvdp = 0.0
				dvdt = 0.0
				cp = 0.0
				t = locale.atof(root.find(".//Temperature").text)+273.15
				if phase_name == 'System':
					for phase in self.phase_names_a:
						if root.find(".//System/Phase[@Type='" + phase + "']/Mass") != None:
							dvdp += float(root.find(".//System/Phase[@Type='" + phase + "']/DvDp").text)
							dvdt += float(root.find(".//System/Phase[@Type='" + phase + "']/DvDt").text)
							cp += float(root.find(".//System/Phase[@Type='" + phase + "']/HeatCapacity").text)
					value = -dvdt/(cp*dvdp + t*dvdt*dvdt)
				else:
					if root.find(".//System/Phase[@Type='" + phase_name + "']/Mass") != None:
						dvdp = float(root.find(".//System/Phase[@Type='" + phase_name + "']/DvDp").text)
						dvdt = float(root.find(".//System/Phase[@Type='" + phase_name + "']/DvDt").text)
						cp = float(root.find(".//System/Phase[@Type='" + phase_name + "']/HeatCapacity").text)
						value = -dvdt/(cp*dvdp + t*dvdt*dvdt)
					else:
						value = 0.0
		else:
			value = 0.0

		return value

	def get_dictionary_of_affinities(self, root, sort=True):
		"""Returns an ordered dictionary of tuples of chemical affinity and phase formulae for
		undersaturated phases in the system.

		Parameters
		----------
		root : type xml.etree.ElementTree
			An xml document tree returned by the function equilibrate_xx
		sort : boolean
			A flag when set to sort the dictionary in order of ascending affinities

		Returns
		-------
		dict : OrderedDict
			A Python ordered dictionary. Dictionary keys are strings naming the phases not in the equilibrium
			assemblage but known to the system.  These phases are by definition undersaturated.
			Dictionary values are tuples consisting of a float value and a string: (affinity, formula).
			For a solution phase the formula is the composition closest to equilibrium with the reported phase
			assemblage.  Dictionary ordering corresponds to array order in get_phase_names(), unless ``sorted``
			is set to True; then entries are ordered by ascending chemical affinity.

		"""
		dict = OrderedDict()
		for phase in self.phase_names_a:
			if root.find(".//Potential/Phase[@Type='" + phase + "']") != None:
				affinity = locale.atof(root.find(".//Potential/Phase[@Type='" + phase + "']/Affinity").text)
				formulae = root.find(".//Potential/Phase[@Type='" + phase + "']/Formula").text
				if affinity == 0.0:
					affinity = 999999.0
				dict[phase] = (affinity, formulae)
		if sort == True:
			return OrderedDict(sorted(dict.items(), key=lambda t:t[1]))
		else:
			return dict

	def output_summary(self, root, printT=True, printP=True, printMass=False, printSysWt=False, printSysM=False, printPhs=True, printPhsWt=False, printPhsM=False):
		"""Prints information about the specified equilibrium phase assemblage.

		Parameters
		----------
		root : type xml.etree.ElementTree
			An xml document tree returned by the function equilibrate_tp
		printT : bool, optional, default=True
			Print the system temperature in degrees centigrade
		printP : bool, optional, default=True
			Print the system pressure in mega-Pascals
		printMass ; bool, optional, default=False
			Print the mass of the system in grams.
		printSysWt : bool, optional, default=False
			Print the compoistion of the system in wt% oxides
		printSysM : bool, optional, default=False
			Print the composition of the system in moles of elements
		printPhs : bool, optional, default=True
			Print the phases present in the system, their masses (in grams) and their chemical formulas
		printPhsWt : bool, optional, default=False
			Print the composition of each phase in wt% oxides (most often used in conjunction with printPhs=True)
		printPhsM : bool, optional, default=False
			Print the composition of each phase in moles of end-member components (most often used in conjunction with printPhs=True)

		"""
		if printT:
			print ("{0:<10s} {1:>8.2f}".format("T (°C)", locale.atof(root.find(".//Temperature").text)))
		if printP:
			print ("{0:<10s} {1:>8.2f}".format("P (MPa)", locale.atof(root.find(".//Pressure").text)*1000.0))
		if printMass:
			print ("{0:<10s} {1:>8.3f}".format("Mass (g)", float(root.find(".//Mass").text)))

		if printSysM:
			print ("Bulk composition in elemental abundances (moles):")
			bcElements = list(root.findall(".//Composition/Element"))
			for element in bcElements:
				print ("   {0:>2s} {1:>8.5f}".format(element.attrib['Type'], float(element.text)))

		if printSysWt:
			print ("Bulk composition in oxide abundances (wt %):")
			bcOxides = list(root.findall(".//Composition/Oxide"))
			for oxide in bcOxides:
				print ("   {0:>5s} {1:>8.4f}".format(oxide.attrib['Type'], float(oxide.text)))

		phases = list(root.findall(".//System/Phase"))
		for phase in phases:
			if printPhs:
				formula = phase.find("Formula").text.split(" ")
				if len(formula) == 1:
					print ("{0:<15.15s} {1:>8.4f} (g)  {2:<60s}".format(phase.attrib['Type'], float(phase.find("Mass").text), formula[0]))
				else:
					# this formula consists of oxide, value pairs (liquid phase)
					n_oxides = len(formula)/2
					n_oxides_1 = 2*int(n_oxides/4)
					n_oxides_2 = int(n_oxides - n_oxides_1)
					line_1 = ""
					for i in range(n_oxides_1):
						line_1 = line_1 + ' ' + formula[2*i] + ' ' + formula[2*i+1]
					print ("{0:<15.15s} {1:>8.4f} (g)  {2:<60s}".format(phase.attrib['Type'], float(phase.find("Mass").text), line_1))
					line_2 = ""
					for i in range(n_oxides_2):
						line_2 = line_2 + ' ' + formula[n_oxides_1*2+2*i] + ' ' + formula[n_oxides_1*2+2*i+1]
					print ("{0:<29.29s}{1:<60s}".format(" ", line_2))
			if printPhsWt:
				oxides = list(phase.findall("Oxide"))
				for oxide in oxides:
					value = float(oxide.text)
					if (value != 0.0):
						print ("   {0:<5s} {1:>8.4f}".format(oxide.attrib['Type'], float(oxide.text)))
			if printPhsM:
				components = list(phase.findall("Component"))
				for component in components:
					value = float(component.text)
					if (value != 0.0):
						print ("   {0:<15s} {1:>9.6f}".format(component.attrib['Name'], float(component.text)))

	# =================================================================
	# Block of Excel workbook functions for output
	# =================================================================

	def start_excel_workbook_with_sheet_name(self, sheetName="Summary"):
		"""Create an Excel workbook with one named sheet.

		Parameters
		----------
		sheetName : string
			Sheet name in the new empty Excel workbook

		Returns
		-------
		wb : type Workbook
		    A pointer to an Excel workbook

		Notes
		-----
		Part of the Excel workbook functions subpackage

		"""
		wb = Workbook()
		ws = wb.active
		ws.title = sheetName
		self.row = 0
		return wb

	def add_sheet_to_workbook_named(self, wb, sheetName):
		"""Creates a new sheet in an existing Excel workbook.

		Parameters
		----------
		wb : type Workbook
			A pointer to an existing Excel workbook
		sheetName : string
			New sheet name in the specified Excel workbook

		Returns
		-------
		ws : type Worksheet
			A pointer to an Excel worksheet in the specified workbook, ``wb``

		Notes
		-----
		Part of the Excel workbook functions subpackage

		"""
		ws = wb.create_sheet(title=sheetName)
		return ws

	def write_to_cell_in_sheet(self, ws, col, row, value, format='general'):
		"""Writes information into the specified row, col on the specified worksheet.

		Parameters
		----------
		ws : type Worksheet
			A pointer to a previously created Excel worksheet (see add_sheet_to_workbook_named)
		col : int
			Column number to write entry to.  Numbering starts at column one.
		row : int
			Row number to write entry to. Numbering starts at row one.
		value : float
			Value to place at entry.
		format : string, optional
			Format to use for value.
			'general' is the default; no formatting applied
			'number' formats as '0.00'
			'scientific' formats as '0.00E+00'

		Notes
		-----
		Part of the Excel workbook functions subpackage

		"""
		if format == 'number':
			ws.cell(column=col, row=row, value=float(value)).number_format = '0.00'
		elif format == 'scientific':
			ws.cell(column=col, row=row, value=float(value)).number_format = '0.00E+00'
		else:
			ws.cell(column=col, row=row, value=value)

	def write_excel_workbook(self, wb, fileName="junk.xlsx"):
		"""Writes the specified Excel workbook to disk.

		Parameters
		----------
		wb : type Workbook
			A pointer to an existing Excel workbook
		fileName : string, optional
			Name of the file that will contain the specified Excel notebook.
			Default file name is 'junk.xlsx'

		Notes
		-----
		Part of the Excel workbook functions subpackage

		"""
		wb.save(filename = fileName)

	def update_excel_workbook(self, wb, root):
		"""Writes the specified equilibrium system state to the specified Excel workbook.

		Parameters
		----------
		wb : type Workbook
			A pointer to an existing Excel workbook
		root : type xml.etree.ElementTree
			An xml document tree returned by the function equilibrate_tp

		Notes
		-----
		The workbook is structured with a Summary worksheet and one worksheet for each equilibrium
		phase.  The evolution of the system is recorded in successive rows.  Row integrity is maintained
		across all sheets. This function may be called repeatedly using different ``root`` objects.
		Part of the Excel workbook functions subpackage.

		"""
		t = locale.atof(root.find(".//Temperature").text)
		p = locale.atof(root.find(".//Pressure").text)*1000.0
		bcElements = list(root.findall(".//Composition/Element"))
		bcOxides = list(root.findall(".//Composition/Oxide"))

		wsSummary = wb.get_sheet_by_name("Summary")
		if (self.row == 0):
			col = 1
			self.row = 1
			self.write_to_cell_in_sheet(wsSummary, col, self.row, "T °C")
			col += 1
			self.write_to_cell_in_sheet(wsSummary, col, self.row, "P MPa")
			col += 1
			self.write_to_cell_in_sheet(wsSummary, col, self.row, "Mass g")
			col += 1
			for element in bcElements:
				self.write_to_cell_in_sheet(wsSummary, col, self.row, element.attrib['Type'])
				col += 1
			for oxide in bcOxides:
				self.write_to_cell_in_sheet(wsSummary, col, self.row, oxide.attrib['Type'])
				col += 1

		self.row += 1
		col = 1
		self.write_to_cell_in_sheet(wsSummary, col, self.row, t, format='number')
		col += 1
		self.write_to_cell_in_sheet(wsSummary, col, self.row, p, format='number')
		col += 1
		self.write_to_cell_in_sheet(wsSummary, col, self.row, root.find(".//Mass").text, format='number')
		col += 1
		for element in bcElements:
			self.write_to_cell_in_sheet(wsSummary, col, self.row, element.text, format='scientific')
			col += 1
		for oxide in bcOxides:
			self.write_to_cell_in_sheet(wsSummary, col, self.row, oxide.text, format='number')
			col += 1

		phases = list(root.findall(".//System/Phase"))
		for phase in phases:
			phaseType = phase.attrib['Type']
			oxides = list(phase.findall("Oxide"))
			components = list(phase.findall("Component"))

			try:
				wsPhase = wb.get_sheet_by_name(phaseType)
			except KeyError:
				wsPhase = wb.create_sheet(phaseType)
				col = 1
				self.write_to_cell_in_sheet(wsPhase, col, 1, "T °C")
				col += 1
				self.write_to_cell_in_sheet(wsPhase, col, 1, "P MPa")
				col += 1
				self.write_to_cell_in_sheet(wsPhase, col, 1, "Mass g")
				col += 1
				self.write_to_cell_in_sheet(wsPhase, col, 1, "Formula")
				col += 1
				for oxide in oxides:
					self.write_to_cell_in_sheet(wsPhase, col, 1, oxide.attrib['Type'])
					col += 1
				for component in components:
					self.write_to_cell_in_sheet(wsPhase, col, 1, component.attrib['Name'])
					col += 1

			col = 1
			self.write_to_cell_in_sheet(wsPhase, col, self.row, t, format='number')
			col += 1
			self.write_to_cell_in_sheet(wsPhase, col, self.row, p, format='number')
			col += 1
			self.write_to_cell_in_sheet(wsPhase, col, self.row, phase.find("Mass").text, format='number')
			col += 1
			self.write_to_cell_in_sheet(wsPhase, col, self.row, phase.find("Formula").text)
			col += 1
			for oxide in oxides:
				self.write_to_cell_in_sheet(wsPhase, col, self.row, oxide.text, format='number')
				col += 1
			for component in components:
				self.write_to_cell_in_sheet(wsPhase, col, self.row, component.text, format='scientific')
				col += 1
